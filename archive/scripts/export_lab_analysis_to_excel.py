#!/usr/bin/env python3
"""
Export specific lab analysis to Excel file.
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Import the analysis function
from analyze_specific_lab import analyze_lab, extract_metrics_from_runtime


def create_excel_for_lab(lab_id: str, server: str, output_file: Path):
    """Create Excel file for specific lab analysis"""
    
    runtime_dir = Path("unified_cache/runtime_reports")
    analysis = analyze_lab(lab_id, server, runtime_dir)
    
    if not analysis.get("all_bots"):
        print(f"❌ No valid backtests found for lab {lab_id} on {server}")
        return False
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Top 5 Bots
    ws1 = wb.active
    ws1.title = "Top 5 Bots"
    
    # Sheet 2: All Bots
    ws2 = wb.create_sheet("All Bots")
    
    # Sheet 3: Statistics
    ws3 = wb.create_sheet("Statistics")
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    lab_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Sheet 1: Top 5
    row = 1
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = f"Lab Analysis: {lab_id} ({server})"
    ws1[f'A{row}'].font = title_font
    ws1[f'A{row}'].alignment = center_align
    row += 1
    
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Bots: {analysis['total_bots']}"
    ws1[f'A{row}'].font = Font(size=10)
    ws1[f'A{row}'].alignment = center_align
    row += 2
    
    headers = ['Rank', 'Backtest ID', 'Initial Balance', 'End Balance', 'ROI %', 'Max Drawdown %', 'Winrate %', 'Total Trades']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    for rank, bot in enumerate(analysis["top_5"], 1):
        ws1.cell(row=row, column=1, value=rank).border = border
        ws1.cell(row=row, column=2, value=bot.get('backtest_id', '')).border = border
        ws1.cell(row=row, column=3, value=bot.get('initial_balance', 0)).number_format = '$#,##0.00'
        ws1.cell(row=row, column=3).alignment = right_align
        ws1.cell(row=row, column=3).border = border
        ws1.cell(row=row, column=4, value=bot.get('end_balance', 0)).number_format = '$#,##0.00'
        ws1.cell(row=row, column=4).alignment = right_align
        ws1.cell(row=row, column=4).border = border
        ws1.cell(row=row, column=5, value=bot.get('roi', 0)/100).number_format = '0.00%'
        ws1.cell(row=row, column=5).alignment = right_align
        ws1.cell(row=row, column=5).border = border
        ws1.cell(row=row, column=6, value=bot.get('max_drawdown', 0)/100).number_format = '0.00%'
        ws1.cell(row=row, column=6).alignment = right_align
        ws1.cell(row=row, column=6).border = border
        ws1.cell(row=row, column=7, value=bot.get('winrate', 0)/100).number_format = '0.00%'
        ws1.cell(row=row, column=7).alignment = right_align
        ws1.cell(row=row, column=7).border = border
        ws1.cell(row=row, column=8, value=bot.get('total_trades', 0)).number_format = '#,##0'
        ws1.cell(row=row, column=8).alignment = right_align
        ws1.cell(row=row, column=8).border = border
        row += 1
    
    # Sheet 2: All Bots
    row = 1
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    for rank, bot in enumerate(analysis["all_bots"], 1):
        ws2.cell(row=row, column=1, value=rank).border = border
        ws2.cell(row=row, column=2, value=bot.get('backtest_id', '')).border = border
        ws2.cell(row=row, column=3, value=bot.get('initial_balance', 0)).number_format = '$#,##0.00'
        ws2.cell(row=row, column=3).alignment = right_align
        ws2.cell(row=row, column=3).border = border
        ws2.cell(row=row, column=4, value=bot.get('end_balance', 0)).number_format = '$#,##0.00'
        ws2.cell(row=row, column=4).alignment = right_align
        ws2.cell(row=row, column=4).border = border
        ws2.cell(row=row, column=5, value=bot.get('roi', 0)/100).number_format = '0.00%'
        ws2.cell(row=row, column=5).alignment = right_align
        ws2.cell(row=row, column=5).border = border
        ws2.cell(row=row, column=6, value=bot.get('max_drawdown', 0)/100).number_format = '0.00%'
        ws2.cell(row=row, column=6).alignment = right_align
        ws2.cell(row=row, column=6).border = border
        ws2.cell(row=row, column=7, value=bot.get('winrate', 0)/100).number_format = '0.00%'
        ws2.cell(row=row, column=7).alignment = right_align
        ws2.cell(row=row, column=7).border = border
        ws2.cell(row=row, column=8, value=bot.get('total_trades', 0)).number_format = '#,##0'
        ws2.cell(row=row, column=8).alignment = right_align
        ws2.cell(row=row, column=8).border = border
        row += 1
    
    # Sheet 3: Statistics
    stats = analysis["statistics"]
    row = 1
    ws3.cell(row=row, column=1, value="Metric").font = Font(bold=True)
    ws3.cell(row=row, column=2, value="Value").font = Font(bold=True)
    row += 1
    
    stats_data = [
        ("Total Bots", analysis["total_bots"]),
        ("Average ROI", f"{stats['avg_roi']:.2f}%"),
        ("Average Winrate", f"{stats['avg_winrate']:.2f}%"),
        ("Average Max Drawdown", f"{stats['avg_max_drawdown']:.2f}%"),
        ("Total Trades (all bots)", f"{stats['total_trades']:,}"),
        ("Best ROI", f"{stats['best_roi']:.2f}%"),
        ("Best Winrate", f"{stats['best_winrate']:.2f}%"),
        ("Worst Max Drawdown", f"{stats['worst_max_dd']:.2f}%"),
    ]
    
    for metric, value in stats_data:
        ws3.cell(row=row, column=1, value=metric)
        ws3.cell(row=row, column=2, value=value)
        row += 1
    
    # Set column widths
    for ws in [ws1, ws2]:
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")
    return True


def main():
    """Main entry point"""
    if len(sys.argv) < 3:
        print("Usage: python export_lab_analysis_to_excel.py <lab_id> <server>")
        print("Example: python export_lab_analysis_to_excel.py ae9637f2-1893-4971-86e5-62cf16e050d2 srv01")
        return 1
    
    lab_id = sys.argv[1]
    server = sys.argv[2]
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = Path(f"lab_{lab_id}_{server}_{timestamp}.xlsx")
    
    create_excel_for_lab(lab_id, server, output_file)
    
    print(f"\n✅ Excel file created: {output_file}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
