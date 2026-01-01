#!/usr/bin/env python3
"""
Export lab analysis results to Excel file.
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def create_excel_from_analysis(json_file: Path, output_file: Path):
    """Create Excel file from analysis JSON"""
    
    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 5 Bots by Lab"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    lab_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Start row
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"Lab Analysis Report - Top 5 Bots Per Lab"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    row += 1
    
    # Summary info
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"Generated: {data['generated']} | Total Labs: {data['total_labs']}"
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = center_align
    row += 2
    
    # Process each lab
    for lab_id, lab_data in sorted(data['labs'].items()):
        # Lab header
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"LAB: {lab_id}"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = lab_header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{row}'].border = border
        row += 1
        
        # Column headers
        headers = ['Rank', 'Backtest ID', 'Initial Balance', 'End Balance', 'ROI %', 'Max Drawdown %', 'Winrate %', 'Total Trades']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Bot data
        for rank, bot in enumerate(lab_data['top_bots'], 1):
            initial_bal = bot.get('initial_balance', 0.0)
            end_bal = bot.get('end_balance', 0.0)
            roi = bot.get('roi_percentage', 0.0)
            max_dd = bot.get('max_drawdown', 0.0)
            winrate = bot.get('winrate', 0.0)
            trades = bot.get('total_trades', 0)
            
            ws.cell(row=row, column=1, value=rank).border = border
            ws.cell(row=row, column=2, value=bot.get('backtest_id', '')).border = border
            ws.cell(row=row, column=3, value=initial_bal).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).alignment = right_align
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=end_bal).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).alignment = right_align
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=5, value=roi).number_format = '0.00%'
            ws.cell(row=row, column=5).alignment = right_align
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=6, value=max_dd).number_format = '0.00%'
            ws.cell(row=row, column=6).alignment = right_align
            ws.cell(row=row, column=6).border = border
            ws.cell(row=row, column=7, value=winrate).number_format = '0.00%'
            ws.cell(row=row, column=7).alignment = right_align
            ws.cell(row=row, column=7).border = border
            ws.cell(row=row, column=8, value=trades).number_format = '#,##0'
            ws.cell(row=row, column=8).alignment = right_align
            ws.cell(row=row, column=8).border = border
            row += 1
        
        # Add summary row for best bot
        row += 1
        best_bot = lab_data['top_bots'][0] if lab_data['top_bots'] else {}
        ws.merge_cells(f'A{row}:H{row}')
        summary_text = f"Best Bot: ROI={best_bot.get('roi_percentage', 0):.2f}% | Winrate={best_bot.get('winrate', 0):.2f}% | Max DD={best_bot.get('max_drawdown', 0):.2f}%"
        ws[f'A{row}'] = summary_text
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 2
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")


def main():
    """Main entry point"""
    # Find latest JSON file
    json_files = list(Path('.').glob('lab_analysis_top_bots_*.json'))
    if not json_files:
        print("ERROR: No analysis JSON file found")
        return 1
    
    latest_json = sorted(json_files)[-1]
    print(f"Using analysis file: {latest_json}")
    
    # Create output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = Path(f"lab_analysis_top_bots_{timestamp}.xlsx")
    
    # Create Excel file
    create_excel_from_analysis(latest_json, output_file)
    
    print(f"\n✅ Excel file created successfully: {output_file}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
